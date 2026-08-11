from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import uuid
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from .config import settings
from .db import connect, now_iso
from .lemonade import LemonadeClient, LemonadeError

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".md", ".csv", ".xlsx"}


def validate_upload_bytes(file_bytes: bytes, filename: str) -> None:
    """Perform conservative format and archive checks before parsing or storing a document."""
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError("PDF, DOCX, TXT, MD, CSV, XLSX файл оруулна уу.")
    if not file_bytes:
        raise ValueError("Хоосон файл оруулах боломжгүй.")
    if len(file_bytes) > settings.upload_max_mb * 1024 * 1024:
        raise ValueError(f"Файл {settings.upload_max_mb}MB-аас их байж болохгүй.")

    if ext == ".pdf":
        if not file_bytes.startswith(b"%PDF-"):
            raise ValueError("PDF файлын бүтэц буруу байна.")
        # We never execute PDF JavaScript and downloads are served as attachments, but
        # rejecting active-action markers further reduces risk from untrusted uploads.
        lowered = file_bytes[: min(len(file_bytes), 8 * 1024 * 1024)].lower()
        if b"/javascript" in lowered or b"/openaction" in lowered or b"/launch" in lowered:
            raise ValueError("Идэвхтэй action/JavaScript агуулсан PDF зөвшөөрөхгүй.")
        return

    if ext in {".docx", ".xlsx"}:
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
                infos = archive.infolist()
                if len(infos) > 5000:
                    raise ValueError("Office файл хэт олон дотоод файлтай байна.")
                total = sum(max(0, item.file_size) for item in infos)
                limit = settings.upload_max_uncompressed_mb * 1024 * 1024
                if total > limit:
                    raise ValueError(f"Office файлын задлагдсан хэмжээ {settings.upload_max_uncompressed_mb}MB-аас их байна.")
                names = {item.filename.replace("\\", "/").lower() for item in infos}
                required = "word/document.xml" if ext == ".docx" else "xl/workbook.xml"
                if "[content_types].xml" not in names or required not in names:
                    raise ValueError("Office файлын бүтэц буруу байна.")
                blocked = ("vbaproject.bin", "/embeddings/", "/oleobjects/", "/activex/")
                if any(any(token in name for token in blocked) for name in names):
                    raise ValueError("Macro эсвэл embedded object агуулсан Office файл зөвшөөрөхгүй.")
        except zipfile.BadZipFile as exc:
            raise ValueError("Office файл эвдэрсэн эсвэл буруу форматтай байна.") from exc
        return

    # Text formats: reject binary-looking payloads.
    sample = file_bytes[:65536]
    if b"\x00" in sample:
        raise ValueError("TXT/CSV/MD нэртэй binary файл зөвшөөрөхгүй.")


@dataclass
class ParsedSection:
    section: str
    text: str


def safe_filename(name: str) -> str:
    base = Path(name).name.strip().replace("\x00", "")
    base = re.sub(r"[^\w.()\- ]+", "_", base, flags=re.UNICODE)
    return base[:180] or "document"


def _clean(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def parse_file(path: Path) -> list[ParsedSection]:
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Дэмжигдээгүй файл: {ext}")
    if ext == ".pdf":
        reader = PdfReader(str(path))
        sections = []
        for index, page in enumerate(reader.pages, 1):
            text = _clean(page.extract_text() or "")
            if text:
                sections.append(ParsedSection(f"Хуудас {index}", text))
        return sections
    if ext == ".docx":
        doc = DocxDocument(str(path))
        sections: list[ParsedSection] = []
        current_title = "Баримт"
        buffer: list[str] = []
        for p in doc.paragraphs:
            text = _clean(p.text)
            if not text:
                continue
            style_name = (p.style.name or "").lower() if p.style else ""
            if style_name.startswith("heading"):
                if buffer:
                    sections.append(ParsedSection(current_title, "\n".join(buffer)))
                    buffer = []
                current_title = text
            else:
                buffer.append(text)
        if buffer:
            sections.append(ParsedSection(current_title, "\n".join(buffer)))
        return sections
    if ext == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        sections = []
        for ws in wb.worksheets:
            lines = []
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if values:
                    lines.append(" | ".join(values))
            text = _clean("\n".join(lines))
            if text:
                sections.append(ParsedSection(f"Sheet: {ws.title}", text))
        return sections
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    if ext == ".csv":
        reader = csv.reader(io.StringIO(text))
        text = "\n".join(" | ".join(cell.strip() for cell in row) for row in reader)
    return [ParsedSection("Баримт", _clean(text))] if _clean(text) else []


def chunk_sections(sections: Iterable[ParsedSection], chunk_size: int = 1300, overlap: int = 180) -> list[ParsedSection]:
    chunks: list[ParsedSection] = []
    for section in sections:
        text = _clean(section.text)
        if not text:
            continue
        if len(text) <= chunk_size:
            chunks.append(section)
            continue
        start = 0
        part = 1
        while start < len(text):
            end = min(len(text), start + chunk_size)
            if end < len(text):
                boundary = max(text.rfind("\n", start, end), text.rfind(". ", start, end))
                if boundary > start + chunk_size // 2:
                    end = boundary + 1
            piece = text[start:end].strip()
            if piece:
                chunks.append(ParsedSection(f"{section.section} · {part}", piece))
            if end >= len(text):
                break
            start = max(start + 1, end - overlap)
            part += 1
    return chunks


def tokenize(text: str) -> list[str]:
    return [token.lower() for token in re.findall(r"[\w%₮$]+", text, flags=re.UNICODE) if len(token) > 1]


def lexical_score(query: str, text: str) -> float:
    q = tokenize(query)
    if not q:
        return 0.0
    haystack = text.lower()
    tokens = tokenize(text)
    if not tokens:
        return 0.0
    token_set = set(tokens)
    overlap = sum(1 for item in set(q) if item in token_set) / max(1, len(set(q)))
    phrase_bonus = 0.25 if query.strip().lower() in haystack else 0.0
    partial_bonus = sum(0.04 for item in set(q) if len(item) >= 4 and item in haystack)
    return min(1.0, overlap * 0.75 + phrase_bonus + partial_bonus)


def cosine(a: list[float], b: list[float]) -> float:
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(y * y for y in b))
    if na == 0 or nb == 0:
        return 0.0
    return dot / (na * nb)


async def ingest_document(
    *,
    file_bytes: bytes,
    filename: str,
    title: str,
    category: str,
    visibility: str,
    department_id: int | None,
    version: str,
    effective_from: str,
    effective_to: str,
    user_id: int,
    embeddings_enabled: bool,
) -> dict:
    filename = safe_filename(filename)
    ext = Path(filename).suffix.lower()
    validate_upload_bytes(file_bytes, filename)

    checksum = hashlib.sha256(file_bytes).hexdigest()
    with connect() as conn:
        duplicate = conn.execute(
            "SELECT id,title FROM documents WHERE checksum=? AND status='active'", (checksum,)
        ).fetchone()
    if duplicate:
        raise ValueError(f"Ижил файл аль хэдийн бүртгэлтэй: {duplicate['title']}")

    doc_id = f"doc_{uuid.uuid4().hex[:16]}"
    stored_name = f"{doc_id}{ext}"
    stored_path = settings.documents_dir / stored_name
    stored_path.write_bytes(file_bytes)

    try:
        sections = parse_file(stored_path)
        chunks = chunk_sections(sections)
        if not chunks:
            raise ValueError("Файлаас унших текст олдсонгүй.")

        vectors: list[list[float]] | None = None
        index_mode = "lexical"
        if embeddings_enabled and not settings.mock_mode:
            try:
                client = LemonadeClient()
                vectors = []
                for start in range(0, len(chunks), 12):
                    batch = [chunk.text for chunk in chunks[start:start + 12]]
                    vectors.extend(await client.embeddings(batch))
                if len(vectors) == len(chunks):
                    index_mode = "hybrid"
            except LemonadeError:
                vectors = None

        timestamp = now_iso()
        with connect() as conn:
            conn.execute(
                """
                INSERT INTO documents(
                    id,title,filename,stored_name,mime_type,size_bytes,category,visibility,department_id,
                    version,status,effective_from,effective_to,checksum,index_mode,chunk_count,created_by,created_at,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?, 'active',?,?,?,?,?,?,?,?)
                """,
                (
                    doc_id, title.strip() or filename, filename, stored_name, ext, len(file_bytes),
                    category or "general", visibility, department_id, version or "1.0",
                    effective_from, effective_to, checksum, index_mode, len(chunks), user_id, timestamp, timestamp,
                ),
            )
            for index, chunk in enumerate(chunks):
                chunk_id = f"chunk_{uuid.uuid4().hex[:18]}"
                vector = vectors[index] if vectors else None
                conn.execute(
                    "INSERT INTO document_chunks(id,document_id,chunk_index,section,content,embedding_json,created_at) VALUES(?,?,?,?,?,?,?)",
                    (
                        chunk_id, doc_id, index, chunk.section, chunk.text,
                        json.dumps(vector) if vector else None, timestamp,
                    ),
                )
        return {"id": doc_id, "title": title or filename, "chunks": len(chunks), "index_mode": index_mode}
    except Exception:
        stored_path.unlink(missing_ok=True)
        raise


async def reindex_document(document_id: str, embeddings_enabled: bool) -> dict:
    with connect() as conn:
        document = conn.execute("SELECT * FROM documents WHERE id=?", (document_id,)).fetchone()
    if not document:
        raise ValueError("Баримт олдсонгүй.")
    path = settings.documents_dir / document["stored_name"]
    sections = parse_file(path)
    chunks = chunk_sections(sections)
    vectors: list[list[float]] | None = None
    index_mode = "lexical"
    if embeddings_enabled and not settings.mock_mode:
        try:
            vectors = []
            client = LemonadeClient()
            for start in range(0, len(chunks), 12):
                vectors.extend(await client.embeddings([c.text for c in chunks[start:start + 12]]))
            if len(vectors) == len(chunks):
                index_mode = "hybrid"
        except LemonadeError:
            vectors = None
    timestamp = now_iso()
    with connect() as conn:
        conn.execute("DELETE FROM document_chunks WHERE document_id=?", (document_id,))
        for index, chunk in enumerate(chunks):
            conn.execute(
                "INSERT INTO document_chunks(id,document_id,chunk_index,section,content,embedding_json,created_at) VALUES(?,?,?,?,?,?,?)",
                (
                    f"chunk_{uuid.uuid4().hex[:18]}", document_id, index, chunk.section, chunk.text,
                    json.dumps(vectors[index]) if vectors else None, timestamp,
                ),
            )
        conn.execute(
            "UPDATE documents SET chunk_count=?, index_mode=?, updated_at=? WHERE id=?",
            (len(chunks), index_mode, timestamp, document_id),
        )
    return {"id": document_id, "chunks": len(chunks), "index_mode": index_mode}
